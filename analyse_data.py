from operations_on_companies_list import first_company_in_list_cell, last_company_in_list_cell, \
    to_fixed, find_company_in_list, last_company_in_hist_list_cell
import os
from highlight_data import color_fundamental_parameters_of_companies_in_list, \
                           color_params_of_champ_list_5years_dividends_increase_in_row, \
                           color_champ_list_after_year_by_year_div_growth_analysis
from openpyxl import load_workbook
from evaluation_properties import StandardEvaluationModel


def get_champ_list_after_fundamental_analysis(champ_worksheet, eval_model):
    """Analysing preliminary list of potential champs in stock list using fundamental analysis parameters"""

    ws = champ_worksheet
    resulted_champ_list = []
    company_list_start_indx = first_company_in_list_cell(ws)
    company_list_end_indx = last_company_in_list_cell(ws)

    for i in range(company_list_start_indx, company_list_end_indx):

        if (  # Div Years
                ws['E' + str(i)].value >= eval_model.div_years
                # Overall AVG divs
                and float(ws['J' + str(i)].value) >= eval_model.avg_divs_overall
                # MR last dividends inc%
                and float(ws['R' + str(i)].value) >= eval_model.mr_last_div_incr
                # EPS a part of profit to dividends
                and ws['Z' + str(i)].value != 'n/a'
                and float(ws['Z' + str(i)].value) < eval_model.eps
                # P/E AVG
                and float(ws['AA' + str(i)].value) <= eval_model.pe_avg
                # MktCap, $Mil
                and float(ws['AL' + str(i)].value) >= eval_model.cap_mil_dollrs
                # Est. div in 5 years Payback, %
                and float(ws['AX' + str(i)].value) >= eval_model.est_div_paybacks_5years_predicted
        ):
            champion = dict()
            champion['company_name'] = ws['A' + str(i)].value
            champion['ticker'] = ws['B' + str(i)].value
            champion['div_years_row'] = ws['E' + str(i)].value
            champion['dividends_avg'] = to_fixed(ws['J' + str(i)].value, 2)
            champion['MR%'] = to_fixed(ws['R' + str(i)].value, 2)
            champion['EPS'] = to_fixed(ws['Z' + str(i)].value, 2)
            champion['AVG_PE'] = to_fixed(ws['AA' + str(i)].value, 2)
            champion['capitalization_mil$'] = to_fixed(ws['AL' + str(i)].value, 2)
            champion['est_divs'] = to_fixed(ws['AX' + str(i)].value, 2)
            resulted_champ_list.append(champion)

    return resulted_champ_list


def get_champ_list_after_5years_dividends_increase_in_row_analysis(input_champ_list, eval_model, hist_worksheet):
    """Analysing dividends magnification for 5 years in row"""

    ws_hist = hist_worksheet
    company_hist_list_start_indx = first_company_in_list_cell(ws_hist)
    company_hist_list_end_indx = last_company_in_hist_list_cell(ws_hist)

    resulted_champ_list_div_row = []
    for company in range(0, len(input_champ_list)):
        i = find_company_in_list(input_champ_list[company]['company_name'], ws_hist, company_hist_list_start_indx,
                                 company_hist_list_end_indx)

        if (
                ws_hist['AB' + str(i)].value >= eval_model.percentage_increase_by_year
                and ws_hist['AC' + str(i)].value >= eval_model.percentage_increase_by_year
                and ws_hist['AD' + str(i)].value >= eval_model.percentage_increase_by_year
                and ws_hist['AE' + str(i)].value >= eval_model.percentage_increase_by_year
                and ws_hist['AF' + str(i)].value >= eval_model.percentage_increase_by_year
        ):
            resulted_champ_list_div_row.append(input_champ_list[company])

    return resulted_champ_list_div_row


# Year by year dividend growth
def get_final_champ_list_after_year_by_year_div_growth_analysis(input_champ_list, hist_worksheet):
    """ Analyse parameters of dividends growth year by year is companies history"""

    ws_hist = hist_worksheet
    company_hist_list_start_indx = first_company_in_list_cell(ws_hist)
    company_hist_list_end_indx = last_company_in_hist_list_cell(ws_hist)
    champ_array_year_by_year_div_growth = []
    for company in range(0, len(input_champ_list)):
        i = find_company_in_list(input_champ_list[company]['company_name'], ws_hist, company_hist_list_start_indx,
                                 company_hist_list_end_indx)

        if (
                ws_hist['C' + str(i)].font.color is None
                and ws_hist['D' + str(i)].font.color is None
                and ws_hist['E' + str(i)].font.color is None
                and ws_hist['F' + str(i)].font.color is None
                and ws_hist['G' + str(i)].font.color is None
                and ws_hist['H' + str(i)].font.color is None
                and ws_hist['I' + str(i)].font.color is None
                #and ws_hist['K' + str(i)].font.color is None
                and ws_hist['L' + str(i)].font.color is None
                #and ws_hist['N' + str(i)].font.color is None
                and ws_hist['O' + str(i)].font.color is None
                and ws_hist['P' + str(i)].font.color is None
                and ws_hist['Q' + str(i)].font.color is None
                and ws_hist['R' + str(i)].font.color is None
                and ws_hist['S' + str(i)].font.color is None
                and ws_hist['T' + str(i)].font.color is None
                and ws_hist['U' + str(i)].font.color is None
                and ws_hist['V' + str(i)].font.color is None
                and ws_hist['W' + str(i)].font.color is None
                and ws_hist['X' + str(i)].font.color is None
                and ws_hist['Y' + str(i)].font.color is None
        ):
            champ_array_year_by_year_div_growth.append(input_champ_list[company])

    return champ_array_year_by_year_div_growth


# Year by year dividend growth of Contenders
def get_final_champ_list_after_year_by_year_div_growth_analysis_contenders(input_champ_list, hist_worksheet):
    """ Analyse parameters of dividends growth year by year is companies history"""

    ws_hist = hist_worksheet
    company_hist_list_start_indx = first_company_in_list_cell(ws_hist)
    company_hist_list_end_indx = last_company_in_hist_list_cell(ws_hist)
    champ_array_year_by_year_div_growth = []
    for company in range(0, len(input_champ_list)):
        i = find_company_in_list(input_champ_list[company]['company_name'], ws_hist, company_hist_list_start_indx,
                                 company_hist_list_end_indx)

        if (
                ws_hist['C' + str(i)].font.color is None
                and ws_hist['D' + str(i)].font.color is None
                and ws_hist['E' + str(i)].font.color is None
                and ws_hist['F' + str(i)].font.color is None
                and ws_hist['G' + str(i)].font.color is None
                and ws_hist['H' + str(i)].font.color is None
                and ws_hist['I' + str(i)].font.color is None
                #and ws_hist['K' + str(i)].font.color is None
                and ws_hist['L' + str(i)].font.color is None
                #and ws_hist['N' + str(i)].font.color is None
                and ws_hist['O' + str(i)].font.color is None
                and ws_hist['P' + str(i)].font.color is None
                and ws_hist['Q' + str(i)].font.color is None
                and ws_hist['R' + str(i)].font.color is None
                and ws_hist['S' + str(i)].font.color is None
        ):
            champ_array_year_by_year_div_growth.append(input_champ_list[company])

    return champ_array_year_by_year_div_growth


def analyze_file(excel_data_file_path, analyzed_excel_data_file_path):
    print("Starting stock analyser")

    #excel_data_file_path = r'C:\\Champions.xlsx'
    #analyzed_excel_data_file_path = 'C:\\Result.xlsx'

    company_types = ['Champions', 'Contenders']
    wb = load_workbook(filename=excel_data_file_path, data_only=True)
    ws_hist = wb['Historical']

    for company_type in company_types:
        ws = wb[company_type]

        stnd_model = StandardEvaluationModel()
        stnd_model.div_years_adjustment(company_type)

        # Colored fill for cells
        final_grad_bg_fill, preliminary_grad_bg_fill = stnd_model.color_adjustment(company_type)

        print("Analysing ", company_type, " list")
        # PRELIM CHAMPIONS LOGIC
        # Get a preliminary list of champions according to fundamental  analysis of companies parameters
        prelim_champs_list = get_champ_list_after_fundamental_analysis(ws, stnd_model)
        # Fill with color all the cells that passed fundamental requirements
        color_fundamental_parameters_of_companies_in_list(prelim_champs_list, ws, preliminary_grad_bg_fill)
        # Color fill prelim chaps that could met conditions of 5 years in row of dividend increase
        color_champ_list_after_year_by_year_div_growth_analysis(prelim_champs_list, ws_hist, preliminary_grad_bg_fill)
        # Color fill prelim chaps that could met conditions of year by year div growth
        color_params_of_champ_list_5years_dividends_increase_in_row(prelim_champs_list, ws_hist,
                                                                    preliminary_grad_bg_fill)

        # POST_PRELIM CHAMPIONS LOGIC (5 YEARS DIV GROWTH)
        # Check if preliminary list of companies are met conditions of 5 years in row of dividend increase
        post_prelim_champs_list = get_champ_list_after_5years_dividends_increase_in_row_analysis(prelim_champs_list,
                                                                                                 stnd_model, ws_hist)
        # Color fill all the cells that met conditions of 5 years in row of dividend increase
        color_params_of_champ_list_5years_dividends_increase_in_row(post_prelim_champs_list, ws_hist,
                                                                    preliminary_grad_bg_fill)

        # FINAL CHAMPIONS LOGIC (YEAR BY YEAR DIV GROWTH)
        # Get a final list of champions according analysis of year by year div growth
        if company_type == 'Champions':
            final_champ_list = get_final_champ_list_after_year_by_year_div_growth_analysis(post_prelim_champs_list,
                                                                                           ws_hist)
        else:
            final_champ_list = get_final_champ_list_after_year_by_year_div_growth_analysis_contenders(
                post_prelim_champs_list,
                ws_hist)
        # Color parameters of year by year div growth with green fill
        color_champ_list_after_year_by_year_div_growth_analysis(final_champ_list, ws_hist, final_grad_bg_fill)
        # Color parameters of year by year div growth with green fill (for final list)
        color_params_of_champ_list_5years_dividends_increase_in_row(final_champ_list, ws_hist, final_grad_bg_fill)
        # Color all the cells of fundamental parameters fo final champ list
        color_fundamental_parameters_of_companies_in_list(final_champ_list, ws, final_grad_bg_fill)

    print("Saving data to ", analyzed_excel_data_file_path)

    wb.save(analyzed_excel_data_file_path)
    print("DONE!")

    os.startfile(analyzed_excel_data_file_path)

"""def binary_search(company_name, worksheet, first_indx, last_indx):
    low = first_indx
    high = last_indx

    while low <= high:
        mid = (low + high)//2
        #Company name
        guess = worksheet['A' + str(mid)].value
        print("Guess ", guess)
        print("company_name ", company_name)
        print("mid ", mid)
        print(guess < company_name)
        if guess == company_name:
            return mid
        if guess > company_name:
            high = mid - 1
        else:
            print("LOW")
            low = mid - 1
            print("LOW = ", low)
    return None
"""
